import json
import queue
import traceback
from tkinter import messagebox
import tkinter as tk
import fitz

# from PDFNetPython3.PDFNetPython import PDFNet, Convert
from apryse_sdk.PDFNetPython import PDFNet, Convert
from openpyxl import Workbook, load_workbook
# import win32con
# import win32gui
from selenium import webdriver
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.chrome.service import Service
from selenium.webdriver import EdgeOptions as Options
from selenium.webdriver import EdgeService as Service
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from subprocess import CREATE_NO_WINDOW
import time
from datetime import datetime, timedelta
from PyPDF2 import PdfReader, PdfWriter
import shutil
from apryse_sdk import *
from openpyxl.styles import Font
from threading import Event, Thread
from queue import Queue

log_event = Event()
status_queue = Queue()

PDFNet.Initialize("demo:1696421591125:7c132c6a03000000008881a049c001583c4c2ee15b2e2f3f0940c126ec")
PDFNet.AddResourceSearchPath(r"Lib\Windows")


def extract_data(pdf_path):
    data = {}
    os.mkdir("htmls")
    os.mkdir("pdfs")
    with open(pdf_path, "rb") as pdf_file:
        reader = PdfReader(pdf_file)
        for i, page in enumerate(reader.pages):
            output = PdfWriter()
            output.add_page(page)
            with open(fr"pdfs\page{i}.pdf", "wb") as pdf_stream:
                output.write(pdf_stream)
    pdf_list = os.listdir(r"pdfs")
    for i, pdf in enumerate(pdf_list):
        Convert.ToExcel(fr"pdfs\{pdf}", fr"htmls\page{i}.xlsx")
    text = ""
    with fitz.open(pdf_path) as doc:
        for page in doc:
            text += "\n" + page.get_text()
    wb = load_workbook(r"htmls\page0.xlsx")
    word = str(wb.active["S3"].value)
    iecbr = anagrams(word, text)
    data["iecbr"] = iecbr
    word = str(wb.active["P2"].value)
    port_code = anagrams(word, text)
    data["port_code"] = port_code
    word = str(wb.active["S2"].value)
    beno = anagrams(word, text)
    data["beno"] = beno
    be_date = wb.active["V2"].value.strftime("%d/%m/%Y")
    data["be_date"] = be_date
    word = str(wb.active["S4"].value)
    gstin_type = anagrams(word, text)
    data["gstin_type"] = gstin_type
    pkg = str(wb.active["S8"].value)
    data["pkg"] = pkg
    g_wt = str(wb.active["AB8"].value)
    data["g_wt"] = g_wt
    word = str(wb.active["E12"].value)
    mode = anagrams(word, text)
    data["mode"] = mode
    word = str(wb.active["G12"].value)
    def_be = anagrams(word, text)
    data["def_be"] = def_be
    word = str(wb.active["Y12"].value)
    hss = anagrams(word, text)
    data["hss"] = hss
    word = str(wb.active["H13"].value)
    country_of_origin = anagrams(word, text)
    data["country_of_origin"] = country_of_origin
    word = str(wb.active["H14"].value)
    port_of_loading = anagrams(word, text)
    data["port_of_loading"] = port_of_loading
    bcd = str(wb.active["C23"].value)
    print(bcd)
    data['bcd'] = bcd
    acd = str(wb.active["F23"].value)
    print(acd)
    data['acd'] = acd
    sws = str(wb.active["I23"].value)
    print(sws)
    data['sws'] = sws
    nccd = str(wb.active["L23"].value)
    print(nccd)
    data['nccd'] = nccd
    add = str(wb.active["N23"].value)
    print(add)
    data['add'] = add
    cvd = str(wb.active["P23"].value)
    data['cvd'] = cvd
    print(cvd)
    igst = str(wb.active["T23"].value)
    print(igst)
    data['igst'] = igst
    g_cess = str(wb.active["AA23"].value)
    print(g_cess)
    data['g_cess'] = g_cess
    tot_ass_val = str(wb.active["AE23"].value)
    print(tot_ass_val)
    data['tot_ass_value'] = tot_ass_val
    total_duty = str(wb.active["P25"].value)
    print(total_duty)
    data['total_duty'] = total_duty
    int_val = str(wb.active["T25"].value)
    print(int_val)
    data['int_val'] = int_val
    pnlty = str(wb.active["X25"].value)
    print(pnlty)
    data['pnlty'] = pnlty
    fine = str(wb.active["AA25"].value)
    print(fine)
    data['fine'] = fine
    tot_amt = str(wb.active["AE25"].value)
    print(tot_amt)
    data['tot_amt'] = tot_amt
    word = str(wb.active["P28"].value).split()
    if len(word) == 2:
        mawb = "".join([anagrams(word[0], text), anagrams(word[1], text)])
    else:
        mawb = anagrams(word[0], text)
    print(mawb)
    data['mawb'] = mawb
    m_date = wb.active["S28"].value.strftime("%d/%m/%Y")
    data['m_date'] = m_date
    print(m_date)
    word = str(wb.active["W28"].value).split()
    if len(word) == 2:
        hawb = "".join([anagrams(word[0], text), anagrams(word[1], text)])
    else:
        hawb = anagrams(word[0], text)
    print(hawb)
    data['hawb'] = hawb
    h_date = wb.active["AB28"].value.strftime("%d/%m/%Y")
    print(h_date)
    print("hdate")
    data['h_date'] = h_date
    initial = 1
    while "".join(sorted("1.BOND NO.")) not in "".join(sorted(str(wb.active[f'C{initial}'].value))):
        initial += 1
    initial += 1
    bond_no = str(wb.active[f"C{initial}"].value)
    print(bond_no)
    data['bond_no'] = bond_no
    word = str(wb.active[f"F{initial}"].value)
    port = anagrams(word, text)
    print(port)
    data['port'] = port
    word = str(wb.active[f"I{initial}"].value)
    bond_cd = anagrams(word, text)
    print(bond_cd)
    data['bond_cd'] = bond_cd
    debt_amt = str(wb.active[f"K{initial}"].value)
    print(debt_amt)
    data['debt_amt'] = debt_amt
    bg_amt = str(wb.active[f"N{initial}"].value)
    print(bg_amt)
    data['bg_amt'] = bg_amt
    row = 0
    for j in range(1, wb.active.max_row):
        for k in range(1, wb.active.max_column):
            if "".join(sorted("Examination")) in "".join(sorted(str(wb.active.cell(j, k).value))):
                row = j
                break
    examination = None
    for i in range(1, wb.active.max_column):
        try:
            examination = wb.active.cell(row, i).value.strftime("%d-%b-%Y").upper()
            break
        except:
            continue
    data['examination'] = str(examination)
    row = 0
    for j in range(1, wb.active.max_row):
        for k in range(1, wb.active.max_column):
            if "".join(sorted("OOC DATE")) in "".join(sorted(str(wb.active.cell(j, k).value))):
                row = j
                break
    oocdate = None
    for i in range(1, wb.active.max_column):
        try:
            oocdate = wb.active.cell(row, i).value.strftime("%d-%b-%Y").upper()
            break
        except:
            continue
    data['oocdate'] = str(oocdate)
    column = 0
    row = 0
    for j in range(1, wb.active.max_row):
        for k in range(1, wb.active.max_column):
            if "".join(sorted("2.LCL/ FCL")) in "".join(sorted(str(wb.active.cell(j, k).value))):
                column = k
                row = j
                break
    lcl_fcl = str(wb.active.cell(row + 1, column).value)
    print(lcl_fcl)
    data['lcl_fcl'] = lcl_fcl
    for j in range(1, wb.active.max_row):
        for k in range(1, wb.active.max_column):
            if "".join(sorted("5.CONTAINER NUMBER")) in "".join(sorted(str(wb.active.cell(j, k).value))):
                column = k
                row = j
                break
    word = str(wb.active.cell(row + 1, column).value)
    container_number = anagrams(word, text)
    print(container_number)
    data['container_number'] = container_number
    inv_details = []
    page_list = os.listdir("htmls")
    for page in page_list:
        wbi = load_workbook(fr"htmls\{page}")
        if "".join(sorted("2.FERIGHT")) not in "".join(sorted(str(wbi.active['E27'].value))):
            continue
        print(page)
        detail = {
            "inv_val": str(wbi.active["C28"].value),
            "freight": str(wbi.active["E28"].value),
            "insurance": str(wbi.active["F28"].value * 100) + "%",
            "hss_val": str(wbi.active["H28"].value),
            "term": anagrams(str(wbi.active["D30"].value), text)
        }
        print(detail)
        inv_details.append(detail)
    initial = 1
    while "".join(sorted("2.IIVONCE NO")) not in "".join(sorted(str(wb.active[f'T{initial}'].value))):
        initial += 1
    initial += 1
    i = 0
    inv_list = []
    while wb.active[f"Q{initial}"].value and len(str(wb.active[f"Q{initial}"].value).strip()) > 0:
        word = str(wb.active[f"T{initial}"].value)
        inv_no = anagrams(word, text)
        inv_amt = str(wb.active[f"AC{initial}"].value)
        curr = str(wb.active[f"AH{initial}"].value)
        # curr = anagrams(word, text)
        # print(f"{inv_no}, {inv_amt}, {curr}")
        inv_list.append((inv_no, inv_amt, curr, inv_details[i]["inv_val"], inv_details[i]["freight"], inv_details[i]["insurance"], inv_details[i]["hss_val"], inv_details[i]["term"]))
        initial += 1
        i += 1
    print(inv_list)
    data['inv_list'] = inv_list
    initial = 1
    while "".join(sorted("OOC NO")) not in "".join(sorted(str(wb.active[f'P{initial}'].value))):
        initial += 1
    ooc_no = str(wb.active[f"T{initial}"].value)
    print(ooc_no)
    data['ooc_no'] = ooc_no
    wb.close()
    page_list = os.listdir("htmls")
    for page in page_list:
        wb = load_workbook(fr"htmls\{page}")
        if "".join(sorted("F. LICENCE DETAILS")) in "".join(sorted(str(wb.active['C26'].value))):
            print(page)
            break
    initial = 28
    licence_list = []
    while "".join(sorted("G. CERTIFICATE DETAILS")) not in "".join(sorted(str(wb.active[f'C{initial}'].value))):
        # print(wb.active[f"C{initial}"].value)
        if wb.active[f"C{initial}"].value:
            line = {
                "invsno": str(wb.active[f"C{initial}"].value),
                "itmsno": str(wb.active[f"D{initial}"].value),
                "licslno": str(wb.active[f"E{initial}"].value),
                "licno": str(wb.active[f"H{initial}"].value),
                "licdate": wb.active[f"M{initial}"].value.strftime("%d-%b-%Y"),
                "code": str(wb.active[f"R{initial}"].value),
                "port": anagrams(str(wb.active[f"X{initial}"].value), text),
                "debit_val": str(wb.active[f"AB{initial}"].value),
                "qty": str(wb.active[f"AK{initial}"].value),
                "uqc": str(wb.active[f"AQ{initial}"].value),
                "debit_duty": str(wb.active[f"AV{initial}"].value),
            }
            licence_list.append(line)
        initial += 1
    data["license_details"] = licence_list
    date = wb.active[f"I{initial + 2}"].value
    if date:
        date = date.strftime("%d-%b-%Y")
    data["certificate_details"] = {
        "certificate_number": anagrams(str(wb.active[f"C{initial + 2}"].value), text),
        "date": str(date),
        "type": str(wb.active[f"Q{initial + 2}"].value),
        "prc_level": str(wb.active[f"W{initial + 2}"].value),
        "iec": str(wb.active[f"AD{initial + 2}"].value),
        "branch_sl_no": str(wb.active[f"AN{initial + 2}"].value),
    }
    data["duties"] = []
    if not data["certificate_details"]["certificate_number"]:
        shutil.rmtree("pdfs")
        shutil.rmtree("htmls")
        return data
    page_list = os.listdir("htmls")
    duties = {}
    for page in page_list:
        wb = load_workbook(fr"htmls\{page}")
        if "".join(sorted("PART - III - DUTIES")) in "".join(sorted(str(wb.active['A10'].value))):
            print(page)
            index = 11
            duties[page] = []
            while "".join(sorted("GLOSSARY")) not in "".join(sorted(str(wb.active[f'A{index}'].value))):
                if "".join(sorted("5.ITEM DESCRIPTION")) in "".join(sorted(str(wb.active[f'G{index}'].value))):
                    duties[page].append(index)
                index += 1
    print(duties)
    for key in duties.keys():
        wb = load_workbook(fr"htmls\{key}")
        for i in duties[key]:
            duty = {
                "inv_sn": str(wb.active[f"C{i + 1}"].value),
                "item_sn": str(wb.active[f"D{i + 1}"].value),
                "cth": str(wb.active[f"E{i + 1}"].value),
                "item_desc": " ".join([anagrams(word, text) for word in str(wb.active[f"G{i + 1}"].value).split()]),
                "c_qty": str(wb.active[f"E{i + 3}"].value),
                "s_qty": str(wb.active[f"G{i + 3}"].value),
                "s_uqc": str(wb.active[f"H{i + 3}"].value),
                "assess_value": str(wb.active[f"N{i + 5}"].value),
                "total_duty": str(wb.active[f"V{i + 5}"].value),
                "notn_no_bcd": anagrams(str(wb.active[f"D{i + 7}"].value), text),
                "notn_no_igst": anagrams(str(wb.active[f"I{i + 7}"].value), text),
                "notn_no_g_cess": anagrams(str(wb.active[f"L{i + 7}"].value), text),
                "notn_sno_bcd": anagrams(str(wb.active[f"D{i + 8}"].value), text),
                "rate_bcd": str(wb.active[f"D{i + 9}"].value),
                "amount_bcd": str(wb.active[f"D{i + 10}"].value),
                "amount_igst": str(wb.active[f"I{i + 10}"].value),
                "duty_fg_bcd": str(wb.active[f"D{i + 11}"].value)
            }
            data["duties"].append(duty)
            print(duty)
    shutil.rmtree("pdfs")
    shutil.rmtree("htmls")
    # print(data)
    return data


def anagrams(word, text):
    word_len = len(word)
    for i in range(0, len(text) - word_len + 1):
        textlet = text[i:i + word_len]
        ip = "".join(sorted(textlet))
        op = "".join(sorted(word))
        if ip == op:
            return textlet
    return False


def create_log_book(log_path):
    wb = Workbook()
    wb.active.title = "First Copy"
    wb.create_sheet(title="Final Copy")
    wb.create_sheet(title="Gate Pass")
    wb.create_sheet(title="Licence Details")
    wb.create_sheet(title="Certificate Details")
    wb.create_sheet(title="Duties")
    first_copy_logs = wb["First Copy"]
    first_copy_logs["A1"] = "Sl. No."
    first_copy_logs["A1"].font = Font(bold=True)
    first_copy_logs["B1"] = "Subject"
    first_copy_logs["B1"].font = Font(bold=True)
    first_copy_logs["C1"] = "File Name"
    first_copy_logs["C1"].font = Font(bold=True)
    first_copy_logs["D1"] = "Downloaded"
    first_copy_logs["D1"].font = Font(bold=True)
    first_copy_logs["E1"] = "Uploaded"
    first_copy_logs["E1"].font = Font(bold=True)
    gate_pass_logs = wb["Gate Pass"]
    gate_pass_logs["A1"] = "Sl. No."
    gate_pass_logs["A1"].font = Font(bold=True)
    gate_pass_logs["B1"] = "Subject"
    gate_pass_logs["B1"].font = Font(bold=True)
    gate_pass_logs["C1"] = "File Name"
    gate_pass_logs["C1"].font = Font(bold=True)
    gate_pass_logs["D1"] = "Downloaded"
    gate_pass_logs["D1"].font = Font(bold=True)
    gate_pass_logs["E1"] = "Uploaded"
    gate_pass_logs["E1"].font = Font(bold=True)
    final_copy_logs = wb["Final Copy"]
    final_copy_logs["A1"] = "Sl. No."
    final_copy_logs["A1"].font = Font(bold=True)
    final_copy_logs["B1"] = "Downloaded"
    final_copy_logs["B1"].font = Font(bold=True)
    final_copy_logs["C1"] = "Uploaded"
    final_copy_logs["C1"].font = Font(bold=True)
    final_copy_logs["D1"] = "Subject"
    final_copy_logs["D1"].font = Font(bold=True)
    final_copy_logs["E1"] = "File Name"
    final_copy_logs["E1"].font = Font(bold=True)
    final_copy_logs["F1"] = "Renamed To"
    final_copy_logs["F1"].font = Font(bold=True)
    final_copy_logs["G1"] = "PortCode"
    final_copy_logs["G1"].font = Font(bold=True)
    final_copy_logs["H1"] = "BE No."
    final_copy_logs["H1"].font = Font(bold=True)
    final_copy_logs["I1"] = "BE Date."
    final_copy_logs["I1"].font = Font(bold=True)
    final_copy_logs["J1"] = "IEC/Br"
    final_copy_logs["J1"].font = Font(bold=True)
    final_copy_logs["K1"] = "GSTIN/TYPE"
    final_copy_logs["K1"].font = Font(bold=True)
    final_copy_logs["L1"] = "PKG"
    final_copy_logs["L1"].font = Font(bold=True)
    final_copy_logs["M1"] = "G.WT"
    final_copy_logs["M1"].font = Font(bold=True)
    final_copy_logs["N1"] = "Mode"
    final_copy_logs["N1"].font = Font(bold=True)
    final_copy_logs["O1"] = "DEF BE"
    final_copy_logs["O1"].font = Font(bold=True)
    final_copy_logs["P1"] = "HSS Status"
    final_copy_logs["P1"].font = Font(bold=True)
    final_copy_logs["Q1"] = "Country of Origin"
    final_copy_logs["Q1"].font = Font(bold=True)
    final_copy_logs["R1"] = "Port of Loading"
    final_copy_logs["R1"].font = Font(bold=True)
    final_copy_logs["S1"] = "BCD"
    final_copy_logs["S1"].font = Font(bold=True)
    final_copy_logs["T1"] = "ACD"
    final_copy_logs["T1"].font = Font(bold=True)
    final_copy_logs["U1"] = "SWS"
    final_copy_logs["U1"].font = Font(bold=True)
    final_copy_logs["V1"] = "NCCD"
    final_copy_logs["V1"].font = Font(bold=True)
    final_copy_logs["W1"] = "ADD"
    final_copy_logs["W1"].font = Font(bold=True)
    final_copy_logs["X1"] = "CVD"
    final_copy_logs["X1"].font = Font(bold=True)
    final_copy_logs["Y1"] = "IGST"
    final_copy_logs["Y1"].font = Font(bold=True)
    final_copy_logs["Z1"] = "G. CESS"
    final_copy_logs["Z1"].font = Font(bold=True)
    final_copy_logs["AA1"] = "TOT.ASS VAL"
    final_copy_logs["AA1"].font = Font(bold=True)
    final_copy_logs["AB1"] = "TOTAL DUTY"
    final_copy_logs["AB1"].font = Font(bold=True)
    final_copy_logs["AC1"] = "INT"
    final_copy_logs["AC1"].font = Font(bold=True)
    final_copy_logs["AD1"] = "PNLTY"
    final_copy_logs["AD1"].font = Font(bold=True)
    final_copy_logs["AE1"] = "FINE"
    final_copy_logs["AE1"].font = Font(bold=True)
    final_copy_logs["AF1"] = "TOT. AMOUNT"
    final_copy_logs["AF1"].font = Font(bold=True)
    final_copy_logs["AG1"] = "MAWB NO"
    final_copy_logs["AG1"].font = Font(bold=True)
    final_copy_logs["AH1"] = "M Date"
    final_copy_logs["AH1"].font = Font(bold=True)
    final_copy_logs["AI1"] = "HAWB N)"
    final_copy_logs["AI1"].font = Font(bold=True)
    final_copy_logs["AJ1"] = "H Date"
    final_copy_logs["AJ1"].font = Font(bold=True)
    final_copy_logs["AK1"] = "BOND NO."
    final_copy_logs["AK1"].font = Font(bold=True)
    final_copy_logs["AL1"] = "PORT"
    final_copy_logs["AL1"].font = Font(bold=True)
    final_copy_logs["AM1"] = "BOND CD"
    final_copy_logs["AM1"].font = Font(bold=True)
    final_copy_logs["AN1"] = "DEBT AMT"
    final_copy_logs["AN1"].font = Font(bold=True)
    final_copy_logs["AO1"] = "EXAMINATION"
    final_copy_logs["AO1"].font = Font(bold=True)
    final_copy_logs["AP1"] = "OOC Date"
    final_copy_logs["AP1"].font = Font(bold=True)
    final_copy_logs["AQ1"] = "LCL/FCL"
    final_copy_logs["AQ1"].font = Font(bold=True)
    final_copy_logs["AR1"] = "CONTAINER NUMBER"
    final_copy_logs["AR1"].font = Font(bold=True)
    final_copy_logs["AS1"] = "OOC NO."
    final_copy_logs["AS1"].font = Font(bold=True)
    final_copy_logs["AT1"] = "INVOICE NO."
    final_copy_logs["AT1"].font = Font(bold=True)
    final_copy_logs["AU1"] = "INV. AMT"
    final_copy_logs["AU1"].font = Font(bold=True)
    final_copy_logs["AV1"] = "CURRENCY"
    final_copy_logs["AV1"].font = Font(bold=True)
    final_copy_logs["AW1"] = "INV VAL"
    final_copy_logs["AW1"].font = Font(bold=True)
    final_copy_logs["AX1"] = "FREIGHT"
    final_copy_logs["AX1"].font = Font(bold=True)
    final_copy_logs["AY1"] = "INSURANCE"
    final_copy_logs["AY1"].font = Font(bold=True)
    final_copy_logs["AZ1"] = "HSS"
    final_copy_logs["AZ1"].font = Font(bold=True)
    final_copy_logs["BA1"] = "TERM"
    final_copy_logs["BA1"].font = Font(bold=True)
    licence_detail_logs = wb["Licence Details"]
    licence_detail_logs["A1"] = "BE NO"
    licence_detail_logs["A1"].font = Font(bold=True)
    licence_detail_logs["B1"] = "INVSNO"
    licence_detail_logs["B1"].font = Font(bold=True)
    licence_detail_logs["C1"] = "ITMSNO"
    licence_detail_logs["C1"].font = Font(bold=True)
    licence_detail_logs["D1"] = "LIC SLNO"
    licence_detail_logs["D1"].font = Font(bold=True)
    licence_detail_logs["E1"] = "LIC NO"
    licence_detail_logs["E1"].font = Font(bold=True)
    licence_detail_logs["F1"] = "LIC DATE"
    licence_detail_logs["F1"].font = Font(bold=True)
    licence_detail_logs["G1"] = "CODE"
    licence_detail_logs["G1"].font = Font(bold=True)
    licence_detail_logs["H1"] = "PORT"
    licence_detail_logs["H1"].font = Font(bold=True)
    licence_detail_logs["I1"] = "DEBIT VALUE"
    licence_detail_logs["I1"].font = Font(bold=True)
    licence_detail_logs["J1"] = "QTY"
    licence_detail_logs["J1"].font = Font(bold=True)
    licence_detail_logs["K1"] = "UQC"
    licence_detail_logs["K1"].font = Font(bold=True)
    licence_detail_logs["L1"] = "DEBIT DUTY"
    licence_detail_logs["L1"].font = Font(bold=True)
    certificate_details_logs = wb["Certificate Details"]
    certificate_details_logs["A1"] = "BE NO"
    certificate_details_logs["A1"].font = Font(bold=True)
    certificate_details_logs["B1"] = "Certificate Number"
    certificate_details_logs["B1"].font = Font(bold=True)
    certificate_details_logs["C1"] = "Date"
    certificate_details_logs["C1"].font = Font(bold=True)
    certificate_details_logs["D1"] = "Type"
    certificate_details_logs["D1"].font = Font(bold=True)
    certificate_details_logs["E1"] = "PRC Level"
    certificate_details_logs["E1"].font = Font(bold=True)
    certificate_details_logs["F1"] = "IEC"
    certificate_details_logs["F1"].font = Font(bold=True)
    certificate_details_logs["G1"] = "Branch SLNo"
    certificate_details_logs["G1"].font = Font(bold=True)
    duties_logs = wb["Duties"]
    duties_logs["A1"] = "BE NO"
    duties_logs["A1"].font = Font(bold=True)
    duties_logs["B1"] = "InvSN"
    duties_logs["B1"].font = Font(bold=True)
    duties_logs["C1"] = "ItemSN"
    duties_logs["C1"].font = Font(bold=True)
    duties_logs["D1"] = "CTH"
    duties_logs["D1"].font = Font(bold=True)
    duties_logs["E1"] = "Item Description"
    duties_logs["E1"].font = Font(bold=True)
    duties_logs["F1"] = "C Qty"
    duties_logs["F1"].font = Font(bold=True)
    duties_logs["G1"] = "S Qty"
    duties_logs["G1"].font = Font(bold=True)
    duties_logs["H1"] = "S UQC"
    duties_logs["H1"].font = Font(bold=True)
    duties_logs["I1"] = "Assess Value"
    duties_logs["I1"].font = Font(bold=True)
    duties_logs["J1"] = "Total Duty"
    duties_logs["J1"].font = Font(bold=True)
    duties_logs["K1"] = "Notn No. (BCD)"
    duties_logs["K1"].font = Font(bold=True)
    duties_logs["L1"] = "Notn No. (IGST)"
    duties_logs["L1"].font = Font(bold=True)
    duties_logs["M1"] = "Notn No. (G. CESS)"
    duties_logs["M1"].font = Font(bold=True)
    duties_logs["N1"] = "Notn SNo."
    duties_logs["N1"].font = Font(bold=True)
    duties_logs["O1"] = "Rate"
    duties_logs["O1"].font = Font(bold=True)
    duties_logs["P1"] = "Amount (BCD)"
    duties_logs["P1"].font = Font(bold=True)
    duties_logs["Q1"] = "Amount (IGST)"
    duties_logs["Q1"].font = Font(bold=True)
    duties_logs["R1"] = "Duty Fg"
    duties_logs["R1"].font = Font(bold=True)
    wb.save(log_path)


def extract_name(pdf_path):
    if os.path.exists("htmls"):
        shutil.rmtree("htmls")
    if os.path.exists("pdfs"):
        shutil.rmtree("pdfs")
    os.mkdir("htmls")
    os.mkdir("pdfs")
    with open(pdf_path, "rb") as pdf_file:
        reader = PdfReader(pdf_file)
        output = PdfWriter()
        output.add_page(reader.pages[0])
        with open(r"pdfs\page.pdf", "wb") as pdf_stream:
            output.write(pdf_stream)
    text = ""
    with fitz.open(pdf_path) as doc:
        for page in doc:
            text += "\n" + page.get_text()
    Convert.ToExcel(r"pdfs\page.pdf", r"htmls\page.xlsx")
    wb = load_workbook(r"htmls\page.xlsx")
    be_no = wb.active["S2"].value
    word = str(wb.active["P28"].value).split()
    if len(word) == 2:
        mawb = "".join([anagrams(word[0], text), word[1]])
    else:
        mawb = anagrams(word[0], text)
    word = str(wb.active["W28"].value).split()
    if len(word) == 2:
        hawb = "".join([anagrams(word[0], text), word[1]])
    else:
        hawb = anagrams(word[0], text)
    date = wb.active["S28"].value.strftime("%Y%m%d")
    # print(f"Date: {date}, BE No.: {be_no}, MWAB: {mwab}, HAWB: {hawb}")
    shutil.rmtree("htmls")
    shutil.rmtree("pdfs")
    return f"OOCBOE-{mawb}-{hawb}-{be_no}-{date}.pdf"


def log_first_copy(file_path, subject, file_name):
    wb = load_workbook(file_path)
    ws = wb["First Copy"]
    max_row = ws.max_row
    if max_row == 1:
        ws.cell(max_row + 1, 1).value = 1
    else:
        ws.cell(max_row + 1, 1).value = int(ws.cell(max_row, 1).value) + 1
    ws.cell(max_row + 1, 2).value = subject
    ws.cell(max_row + 1, 3).value = file_name
    ws.cell(max_row + 1, 4).value = datetime.now()
    ws.cell(max_row + 1, 5).value = datetime.now()
    wb.save(file_path)


def log_gate_pass(file_path, subject, file_name):
    wb = load_workbook(file_path)
    ws = wb["Gate Pass"]
    max_row = ws.max_row
    if max_row == 1:
        ws.cell(max_row + 1, 1).value = 1
    else:
        ws.cell(max_row + 1, 1).value = int(ws.cell(max_row, 1).value) + 1
    ws.cell(max_row + 1, 2).value = subject
    ws.cell(max_row + 1, 3).value = file_name
    ws.cell(max_row + 1, 4).value = datetime.now()
    ws.cell(max_row + 1, 5).value = datetime.now()
    wb.save(file_path)


def log_final_copy(file_path, subject, file_name, re_named, data):
    wb = load_workbook(file_path)
    ws = wb["Final Copy"]
    for invoice_no, inv_amt, curr, inv_val, freight, insurance, hss_val, term in data['inv_list']:
        max_row = ws.max_row
        if max_row == 1:
            ws.cell(max_row + 1, 1).value = 1
        else:
            ws.cell(max_row + 1, 1).value = int(ws.cell(max_row, 1).value) + 1
        ws.cell(max_row + 1, 2).value = datetime.now()
        ws.cell(max_row + 1, 3).value = datetime.now()
        ws.cell(max_row + 1, 4).value = subject
        ws.cell(max_row + 1, 5).value = file_name
        ws.cell(max_row + 1, 6).value = re_named
        ws.cell(max_row + 1, 7).value = data['port_code']
        ws.cell(max_row + 1, 8).value = data['beno']
        ws.cell(max_row + 1, 9).value = data['be_date']
        ws.cell(max_row + 1, 10).value = data['iecbr']
        ws.cell(max_row + 1, 11).value = data['gstin_type']
        ws.cell(max_row + 1, 12).value = data['pkg']
        ws.cell(max_row + 1, 13).value = data['g_wt']
        ws.cell(max_row + 1, 14).value = data['mode']
        ws.cell(max_row + 1, 15).value = data['def_be']
        ws.cell(max_row + 1, 16).value = data['hss']
        ws.cell(max_row + 1, 17).value = data['country_of_origin']
        ws.cell(max_row + 1, 18).value = data['port_of_loading']
        ws.cell(max_row + 1, 19).value = data['bcd']
        ws.cell(max_row + 1, 20).value = data['acd']
        ws.cell(max_row + 1, 21).value = data['sws']
        ws.cell(max_row + 1, 22).value = data['nccd']
        ws.cell(max_row + 1, 23).value = data['add']
        ws.cell(max_row + 1, 24).value = data['cvd']
        ws.cell(max_row + 1, 25).value = data['igst']
        ws.cell(max_row + 1, 26).value = data['g_cess']
        ws.cell(max_row + 1, 27).value = data['tot_ass_value']
        ws.cell(max_row + 1, 28).value = data['total_duty']
        ws.cell(max_row + 1, 29).value = data['int_val']
        ws.cell(max_row + 1, 30).value = data['pnlty']
        ws.cell(max_row + 1, 31).value = data['fine']
        ws.cell(max_row + 1, 32).value = data['tot_amt']
        ws.cell(max_row + 1, 33).value = data['mawb']
        ws.cell(max_row + 1, 34).value = data['m_date']
        ws.cell(max_row + 1, 35).value = data['hawb']
        ws.cell(max_row + 1, 36).value = data['h_date']
        ws.cell(max_row + 1, 37).value = data['bond_no']
        ws.cell(max_row + 1, 38).value = data['port']
        ws.cell(max_row + 1, 39).value = data['bond_cd']
        ws.cell(max_row + 1, 40).value = data['debt_amt']
        ws.cell(max_row + 1, 41).value = data['examination']
        ws.cell(max_row + 1, 42).value = data['oocdate']
        ws.cell(max_row + 1, 43).value = data['lcl_fcl']
        ws.cell(max_row + 1, 44).value = data['container_number']
        ws.cell(max_row + 1, 45).value = data['ooc_no']
        ws.cell(max_row + 1, 46).value = invoice_no
        ws.cell(max_row + 1, 47).value = inv_amt
        ws.cell(max_row + 1, 48).value = curr
        ws.cell(max_row + 1, 49).value = inv_val
        ws.cell(max_row + 1, 50).value = freight
        ws.cell(max_row + 1, 51).value = insurance
        ws.cell(max_row + 1, 52).value = hss_val
        ws.cell(max_row + 1, 53).value = term
    ws = wb["Licence Details"]
    for line in data['license_details']:
        max_row = ws.max_row
        ws.cell(max_row + 1, 1).value = data['beno']
        ws.cell(max_row + 1, 2).value = line['invsno']
        ws.cell(max_row + 1, 3).value = line['itmsno']
        ws.cell(max_row + 1, 4).value = line['licslno']
        ws.cell(max_row + 1, 5).value = line['licno']
        ws.cell(max_row + 1, 6).value = line['licdate']
        ws.cell(max_row + 1, 7).value = line['code']
        ws.cell(max_row + 1, 8).value = line['port']
        ws.cell(max_row + 1, 9).value = line['debit_val']
        ws.cell(max_row + 1, 10).value = line['qty']
        ws.cell(max_row + 1, 11).value = line['uqc']
        ws.cell(max_row + 1, 12).value = line['debit_duty']
    ws = wb["Certificate Details"]
    max_row = ws.max_row
    ws.cell(max_row + 1, 1).value = data['beno']
    ws.cell(max_row + 1, 2).value = data["certificate_details"]["certificate_number"]
    ws.cell(max_row + 1, 3).value = data["certificate_details"]["date"]
    ws.cell(max_row + 1, 4).value = data["certificate_details"]["type"]
    ws.cell(max_row + 1, 5).value = data["certificate_details"]["prc_level"]
    ws.cell(max_row + 1, 6).value = data["certificate_details"]["iec"]
    ws.cell(max_row + 1, 7).value = data["certificate_details"]["branch_sl_no"]
    ws = wb["Duties"]
    for duty in data['duties']:
        max_row = ws.max_row
        ws.cell(max_row + 1, 1).value = data['beno']
        ws.cell(max_row + 1, 2).value = duty['inv_sn']
        ws.cell(max_row + 1, 3).value = duty['item_sn']
        ws.cell(max_row + 1, 4).value = duty['cth']
        ws.cell(max_row + 1, 5).value = duty['item_desc']
        ws.cell(max_row + 1, 6).value = duty['c_qty']
        ws.cell(max_row + 1, 7).value = duty['s_qty']
        ws.cell(max_row + 1, 8).value = duty['s_uqc']
        ws.cell(max_row + 1, 9).value = duty['assess_value']
        ws.cell(max_row + 1, 10).value = duty['total_duty']
        ws.cell(max_row + 1, 11).value = duty['notn_no_bcd']
        ws.cell(max_row + 1, 12).value = duty['notn_no_igst']
        ws.cell(max_row + 1, 13).value = duty['notn_no_g_cess']
        ws.cell(max_row + 1, 14).value = duty['notn_sno_bcd']
        ws.cell(max_row + 1, 15).value = duty['rate_bcd']
        ws.cell(max_row + 1, 16).value = duty['amount_bcd']
        ws.cell(max_row + 1, 17).value = duty['amount_igst']
        ws.cell(max_row + 1, 18).value = duty['duty_fg_bcd']
    wb.save(file_path)


def ibm_portal(upload_input):
    credentials = json.loads(open("credentials.json").read())
    service = Service()
    service.creation_flags = CREATE_NO_WINDOW
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-gpu")
    # options.add_experimental_option("detach", True)
    driver = webdriver.Edge(options=options, service=service)
    wait = WebDriverWait(driver, 10)
    status_queue.put("Uploading " + upload_input + " to IBM Portal.")
    driver.get("http://prod.sepl.local/navigator/?desktop=ShahiUser")
    username = wait.until(
        ec.presence_of_element_located((By.ID, 'ecm_widget_layout_NavigatorMainLayout_0_LoginPane_username')))
    username.send_keys(credentials["IBM"]["username"])
    password = driver.find_element(By.ID, 'ecm_widget_layout_NavigatorMainLayout_0_LoginPane_password')
    password.send_keys(credentials["IBM"]["password"])
    login_button = driver.find_element(By.ID, 'ecm_widget_layout_NavigatorMainLayout_0_LoginPane_LoginButton_label')
    login_button.click()
    driver.execute_script("arguments[0].click();", wait.until(ec.presence_of_element_located((By.ID, 'dijit__TreeNode_5_label'))))
    driver.execute_script("arguments[0].click();", wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, 'a[title = "BOE"]'))))
    driver.execute_script("arguments[0].click();", wait.until(ec.presence_of_element_located((By.ID, 'IMPORT_dijit_form_Button_1_label'))))
    file_upload = wait.until(
        ec.presence_of_element_located((By.ID, 'ecm_widget_AddContentItemGeneralPane_0_fileInput')))
    file_upload.send_keys(upload_input)
    time.sleep(2)
    driver.execute_script("arguments[0].click();", driver.find_element(By.ID, "ADD_dijit_form_Button_0"))
    # driver.find_element(By.ID, "ADD_dijit_form_Button_0").click()
    time.sleep(2)


def get_email_list():
    wb = load_workbook("LogEmails.xlsx")
    ws = wb.active
    return "\n".join([ws.cell(i, 1).value for i in range(2, ws.max_row + 1)]) + "\n"


def send_log():
    credentials = json.loads(open("credentials.json").read())
    date = datetime.now()
    date = date - timedelta(days=1)
    DD = str(date.day)
    MM = str(date.month)
    YYYY = str(date.year)
    date_str = "DT-" + YYYY.rjust(4, '0') + "-" + MM.rjust(2, '0') + "-" + DD.rjust(2, '0')
    print(date_str)
    root = r"C:\BOE-automation-" + date_str
    file = root + fr"\BOE-log-{date_str}.xlsx"
    service = Service()
    service.creation_flags = CREATE_NO_WINDOW
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-gpu")
    driver = webdriver.Edge(options=options, service=service)
    driver.get("https://zmail.shahi.co.in:4443/#3")
    wait = WebDriverWait(driver, 10)
    username = wait.until(ec.presence_of_element_located((By.ID, 'username')))
    username.clear()
    username.send_keys(credentials["zimbra"]["username"])
    password = driver.find_element(By.ID, 'password')
    password.clear()
    password.send_keys(credentials["zimbra"]["password"])
    login_button = driver.find_element(By.ID, 'loginButton')
    login_button.click()
    create_mail = wait.until(ec.presence_of_element_located((By.ID, "zb__NEW_MENU_title")))
    create_mail.click()
    attach = wait.until(ec.presence_of_element_located((By.ID, "zb__COMPOSE-1___attachments_btn_title")))
    attach.click()
    time.sleep(1)
    # handle = win32gui.FindWindow(None, "Open")
    # win32gui.PostMessage(handle, win32con.WM_CLOSE, 0, 0)
    upload = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
    upload.send_keys(file)
    to = driver.find_element(By.ID, "zv__COMPOSE-1_to_control")
    to.send_keys(get_email_list())
    subject = driver.find_element(By.ID, "zv__COMPOSE-1_subject_control")
    subject.send_keys(f"BOE Auto email Log details for {date_str}")
    frame = driver.find_element(By.ID, "ZmHtmlEditor1_body_ifr")
    driver.switch_to.frame(frame)
    body = driver.find_element(By.ID, "tinymce")
    body.send_keys(
        f"The Log file for previous days ({date_str}) BOE is in the attachment. This is an automated mail. Do not reply!")
    driver.switch_to.default_content()
    driver.find_element(By.ID, "zb__COMPOSE-1__SEND_title").click()
    time.sleep(2)


def boe_auto_bot():
    credentials = json.loads(open("credentials.json").read())
    date = datetime.now()
    DD = str(date.day)
    MM = str(date.month)
    YYYY = str(date.year)
    date_str = "DT-" + YYYY.rjust(4, '0') + "-" + MM.rjust(2, '0') + "-" + DD.rjust(2, '0')
    print(date_str)
    root = r"C:\BOE-automation-" + date_str
    if not os.path.exists(root):
        os.mkdir(root)
    firstcopy = root + r"\first-copy"
    gatepass = root + r"\gate-pass"
    finalcopy = root + r"\final-copy"
    finalcopyrenamed = root + r"\final-copy-renamed"
    log = os.path.join(root, f'BOE-log-{date_str}.xlsx')
    if not os.path.exists(firstcopy):
        os.mkdir(firstcopy)
    if not os.path.exists(gatepass):
        os.mkdir(gatepass)
    if not os.path.exists(finalcopy):
        os.mkdir(finalcopy)
    if not os.path.exists(finalcopyrenamed):
        os.mkdir(finalcopyrenamed)
    if not os.path.isfile(log):
        create_log_book(log)
    prefs = {"download.default_directory": root}
    service = Service()
    service.creation_flags = CREATE_NO_WINDOW
    options = Options()
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-gpu")
    # options.add_experimental_option("detach", True)
    driver = webdriver.Edge(options=options, service=service)
    driver.get("https://zmail.shahi.co.in:4443/#3")
    wait = WebDriverWait(driver, 10)
    username = wait.until(ec.presence_of_element_located((By.ID, 'username')))
    username.clear()
    username.send_keys(credentials["zimbra"]["username"])
    password = driver.find_element(By.ID, 'password')
    password.clear()
    password.send_keys(credentials["zimbra"]["password"])
    login_button = driver.find_element(By.ID, 'loginButton')
    login_button.click()
    search = wait.until(ec.presence_of_element_located((By.ID, 'zi_search_inputfield')))
    print("Login Complete")
    status_queue.put("Login Complete")
    search.send_keys("BoE is:unread")
    search_btn = wait.until(ec.presence_of_element_located((By.XPATH, '//*[@id="zb__Search__SEARCH_left_icon"]/div')))
    search_btn.click()
    driver.execute_script("arguments[0].click();", wait.until(ec.presence_of_element_located((By.XPATH, '//*[@id="zlha__TV-SR-1__dt"]/div'))))
    time.sleep(2.0)
    email_list = wait.until(ec.presence_of_element_located((By.ID, 'zl__TV-SR-1__rows')))
    menu = driver.find_element(By.ID, 'zb__TV-SR-1__VIEW_MENU_title')
    menu.click()
    panes = driver.find_element(By.ID, 'READING_PANE_2_title')
    panes.click()
    bottom_pane = driver.find_element(By.CSS_SELECTOR, 'td[id *= "bottom__DWT"].ZWidgetTitle')
    bottom_pane.click()
    print("Emails Found!")
    status_queue.put("Email List Found")
    print(email_list.get_attribute('class'))
    emails = email_list.find_elements(By.TAG_NAME, 'li')
    for email in emails:
        subject = email.get_attribute('innerText')
        sub = subject.split('-')[0]
        print(sub)
        status_queue.put("Processing: " + sub)
        open_email = email.find_element(By.CLASS_NAME, 'ZmMsgListColSubject')
        open_email.click()
        time.sleep(2)
        download_link = wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, 'a[title = "Download"]')))
        file_name = driver.find_element(By.CSS_SELECTOR, 'a[title *= ".pdf"')
        file = file_name.get_attribute('title')
        print(file)
        status_queue.put("Downloading: " + file)
        download_link.click()
        time.sleep(5)
        if "first copy" in subject.lower():
            shutil.move(os.path.join(root, file), firstcopy)
            os.rename(os.path.join(firstcopy, file), os.path.join(firstcopy, "FCBOE-" + file))
            ibm_portal(os.path.join(firstcopy, "FCBOE-" + file))
            while log_event.is_set():
                time.sleep(1)
            log_event.set()
            log_first_copy(log, subject, "FCBOE-" + file)
            log_event.clear()
        if "gatepass" in subject.lower():
            shutil.move(os.path.join(root, file), gatepass)
            ibm_portal(os.path.join(gatepass, file))
            while log_event.is_set():
                time.sleep(1)
            log_event.set()
            log_gate_pass(log, subject, file)
            log_event.clear()
        if "final ooc" in subject.lower():
            shutil.copy(os.path.join(root, file), finalcopy)
            shutil.move(os.path.join(root, file), finalcopyrenamed)
            new_file = extract_name(os.path.join(finalcopyrenamed, file))
            data = extract_data(os.path.join(finalcopyrenamed, file))
            os.rename(os.path.join(finalcopyrenamed, file), os.path.join(finalcopyrenamed, new_file))
            ibm_portal(os.path.join(finalcopyrenamed, new_file))
            while log_event.is_set():
                time.sleep(1)
            log_event.set()
            log_final_copy(log, subject, file, new_file, data)
            log_event.clear()


def log_send():
    try:
        send_log()
    except:
        print("Yesterday's Log Couldn't be sent")


def display_error(err, app):
    messagebox.showerror("Error", f"Encountered the following error in {app}:\n{err}")


def run_bot():
    try:
        boe_auto_bot()
    except:
        log_event.clear()
        traceback.print_exc()
        exc_type, exc_value, exc_traceback = sys.exc_info()
        filename = exc_traceback.tb_frame.f_code.co_filename
        lineno = exc_traceback.tb_lineno
        Thread(target=lambda: display_error(traceback.format_exc(), f"file: {filename}, line: {lineno}"), daemon=True).start()
    finally:
        time.sleep(10)
        if os.path.exists("htmls"):
            shutil.rmtree("htmls")
        if os.path.exists("pdfs"):
            shutil.rmtree("pdfs")
        run_bot()


def auto_mail():
    hr = datetime.now().hour
    if 9 <= hr < 10:
        while log_event.is_set():
            time.sleep(1)
        log_event.set()
        log_send()
        log_event.clear()
    time.sleep(3599)
    auto_mail()


def update_label():
    try:
        status = status_queue.get(timeout=5)
        label.config(text=status)
        print(status)
    except queue.Empty:
        print("Nothing Happened")
    finally:
        root.after(1000, update_label)


Thread(target=run_bot, daemon=True).start()
Thread(target=auto_mail, daemon=True).start()
root = tk.Tk()
root.title("BOE Automation")
root.geometry("700x200")
label = tk.Label(root, text="")
label.pack(pady=30)
root.after(1000, update_label)
root.mainloop()

# boe_auto_bot()
